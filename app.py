import os
import re
import json
import hashlib
import threading
from datetime import datetime
from functools import wraps

from flask import Flask, request, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook

STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
app = Flask(__name__, static_folder=STATIC_DIR, static_url_path='')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
SUBMISSIONS_DIR = os.path.join(DATA_DIR, 'submissions')
CONFIG_FILE = os.path.join(DATA_DIR, 'config.json')

PASSWORD_HASH = '240be518fabd2724ddb6f04eeb1da5967448d7e831c08c8fa822809f74c720a9'

HEADERS = ['Timestamp', 'Full Name', 'Phone', 'Age', 'Country', 'Availability', 'Languages']

write_lock = threading.Lock()

os.makedirs(SUBMISSIONS_DIR, exist_ok=True)


def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return jsonify({'result': 'error', 'message': 'Unauthorized'}), 401
        password = auth[7:]
        h = hashlib.sha256(password.encode()).hexdigest()
        if h != PASSWORD_HASH:
            return jsonify({'result': 'error', 'message': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as f:
            return json.load(f)
    return {}


def save_config(data):
    tmp = CONFIG_FILE + '.tmp'
    with open(tmp, 'w') as f:
        json.dump(data, f)
    os.replace(tmp, CONFIG_FILE)


@app.route('/api/submit', methods=['POST'])
def submit():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'result': 'error', 'message': 'Invalid JSON'}), 400

    # Honeypot check
    if data.get('website', '').strip():
        return jsonify({'result': 'success'})

    required = ['fullName', 'phone', 'age', 'country', 'availability']
    for field in required:
        if not data.get(field, '').strip() if isinstance(data.get(field), str) else not data.get(field):
            return jsonify({'result': 'error', 'message': f'Missing field: {field}'}), 400

    today = datetime.now().strftime('%Y-%m-%d')
    filepath = os.path.join(SUBMISSIONS_DIR, f'{today}.xlsx')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    row = [
        timestamp,
        str(data.get('fullName', '')),
        str(data.get('phone', '')),
        str(data.get('age', '')),
        str(data.get('country', '')),
        str(data.get('availability', '')),
        str(data.get('languages', '')),
    ]

    with write_lock:
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
        ws.append(row)
        wb.save(filepath)

    return jsonify({'result': 'success'})


@app.route('/api/config', methods=['GET'])
def get_config():
    return jsonify(load_config())


@app.route('/api/config', methods=['POST'])
@require_auth
def set_config():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'result': 'error', 'message': 'Invalid JSON'}), 400

    config = load_config()
    for key in ('support_url', 'support_icon'):
        if key in data:
            config[key] = data[key]
    save_config(config)
    return jsonify({'result': 'success'})


@app.route('/api/files', methods=['GET'])
@require_auth
def list_files():
    files = []
    for name in sorted(os.listdir(SUBMISSIONS_DIR), reverse=True):
        if not name.endswith('.xlsx'):
            continue
        path = os.path.join(SUBMISSIONS_DIR, name)
        size = os.path.getsize(path)
        try:
            wb = load_workbook(path, read_only=True)
            rows = wb.active.max_row - 1  # exclude header
            wb.close()
        except Exception:
            rows = 0
        files.append({'name': name, 'size': size, 'rows': rows})
    return jsonify({'files': files})


FILENAME_RE = re.compile(r'^\d{4}-\d{2}-\d{2}\.xlsx$')


@app.route('/api/files/<filename>', methods=['GET'])
@require_auth
def download_file(filename):
    if not FILENAME_RE.match(filename):
        return jsonify({'result': 'error', 'message': 'Invalid filename'}), 400
    return send_from_directory(SUBMISSIONS_DIR, filename, as_attachment=True)


@app.route('/api/files/<filename>', methods=['DELETE'])
@require_auth
def delete_file(filename):
    if not FILENAME_RE.match(filename):
        return jsonify({'result': 'error', 'message': 'Invalid filename'}), 400
    path = os.path.join(SUBMISSIONS_DIR, filename)
    if not os.path.exists(path):
        return jsonify({'result': 'error', 'message': 'File not found'}), 404
    os.remove(path)
    return jsonify({'result': 'success'})


@app.route('/')
def index():
    return send_from_directory(STATIC_DIR, 'index.html')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80)
